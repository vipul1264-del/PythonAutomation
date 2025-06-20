# read excel

import openpyxl

from openpyxl import workbook, load_workbook


wb = openpyxl.load_workbook("D:\\datasheet.xlsx")
sheet = wb['data']
print(sheet)
print(sheet['B1'].value)


print(wb.cell(row=6, column=4)).value